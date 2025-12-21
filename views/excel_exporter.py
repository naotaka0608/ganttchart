"""Excel出力機能"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from typing import List
from models import Task


class ExcelExporter:
    """Excelエクスポーター"""

    def __init__(self, tasks: List[Task]):
        self.tasks = tasks
        self.wb = Workbook()

    def export(self, file_path: str):
        """Excelファイルにエクスポート"""
        # タスクリストシートを作成
        self._create_task_list_sheet()

        # ガントチャートシートを作成
        self._create_gantt_chart_sheet()

        # デフォルトシートを削除（最初に作られる空のシート）
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']

        # ファイル保存
        self.wb.save(file_path)

    def _create_task_list_sheet(self):
        """タスクリストシートを作成"""
        ws = self.wb.create_sheet("タスクリスト", 0)

        # ヘッダー作成
        headers = ["タスク名", "進捗率", "担当者", "開始日", "終了日", "期間(日)",
                   "説明", "マイルストーン", "ベースライン開始", "ベースライン終了", "差分(日)"]
        ws.append(headers)

        # ヘッダーのスタイル設定
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # タスクデータを追加（階層構造を表現）
        def add_task_rows(task, level=0):
            """タスクを再帰的に追加"""
            indent = "　" * level
            task_name = f"{indent}{'◆ ' if task.is_milestone else ''}{task.name}"

            row_data = [
                task_name,
                f"{task.progress}%",
                task.assignee or "",
                task.start_date.strftime("%Y-%m-%d"),
                task.end_date.strftime("%Y-%m-%d"),
                task.duration_days,
                task.description or "",
                "○" if task.is_milestone else "",
                task.baseline_start_date.strftime("%Y-%m-%d") if task.baseline_start_date else "",
                task.baseline_end_date.strftime("%Y-%m-%d") if task.baseline_end_date else "",
                f"{task.end_variance_days:+d}日" if task.has_baseline else ""
            ]
            ws.append(row_data)

            for child in task.children:
                add_task_rows(child, level + 1)

        # 階層構造を構築してルートタスクを抽出
        for task in self.tasks:
            task.children = []
        task_dict = {t.id: t for t in self.tasks}
        root_tasks = []
        for task in self.tasks:
            if task.parent_id and task.parent_id in task_dict:
                task_dict[task.parent_id].add_child(task)
            elif task.parent_id is None:
                root_tasks.append(task)

        # ルートタスクから順に追加
        for task in root_tasks:
            add_task_rows(task)

        # 列幅を調整
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 15
        ws.column_dimensions['K'].width = 12

        # セルの枠線とアライメントを設定
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=11):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1:
                    cell.alignment = Alignment(vertical="center")

    def _create_gantt_chart_sheet(self):
        """ガントチャートシートを作成"""
        ws = self.wb.create_sheet("ガントチャート", 1)

        # すべてのタスクをフラット化
        all_tasks = self._flatten_tasks(self.tasks)

        if not all_tasks:
            return

        # 日付範囲を計算
        dates = []
        for task in all_tasks:
            dates.append(task.start_date)
            dates.append(task.end_date)

        min_date = min(dates)
        max_date = max(dates)

        # 余白を追加
        min_date -= timedelta(days=3)
        max_date += timedelta(days=3)

        # 日付のリストを作成
        date_list = []
        current_date = min_date
        while current_date <= max_date:
            date_list.append(current_date)
            current_date += timedelta(days=1)

        # ヘッダー行（日付）
        header_row = ["タスク名", "進捗率", "担当者"] + [d.strftime("%m/%d") for d in date_list]
        ws.append(header_row)

        # ヘッダーのスタイル
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, cell in enumerate(ws[1], start=1):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            if col_idx >= 4:  # 日付列
                ws.column_dimensions[get_column_letter(col_idx)].width = 3

        # 列幅設定
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 8
        ws.column_dimensions['C'].width = 12

        # タスク行を追加
        row_idx = 2
        for task in all_tasks:
            # タスク名（インデントで階層表現）
            level = self._get_task_level(task, all_tasks)
            indent = "　" * level
            task_name = f"{indent}{'◆ ' if task.is_milestone else ''}{task.name}"

            # タスク情報列
            row_data = [task_name, f"{task.progress}%", task.assignee or ""]

            # 日付列：タスクの期間に応じてマーク
            for d in date_list:
                if task.start_date <= d <= task.end_date:
                    row_data.append("■")
                else:
                    row_data.append("")

            ws.append(row_data)

            # タスクバーのセルに色を付ける
            task_fill = self._get_task_color_fill(task)
            for col_idx, d in enumerate(date_list, start=4):
                if task.start_date <= d <= task.end_date:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = task_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # ベースラインがある場合は下線で表示
            if task.has_baseline:
                baseline_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
                for col_idx, d in enumerate(date_list, start=4):
                    if task.baseline_start_date <= d <= task.baseline_end_date:
                        cell = ws.cell(row=row_idx, column=col_idx)
                        # 既存の塗りつぶしがある場合はそのまま、ない場合はグレー
                        if not (task.start_date <= d <= task.end_date):
                            cell.fill = baseline_fill

            row_idx += 1

        # 今日の日付に縦線（列を強調）
        today = date.today()
        if min_date <= today <= max_date:
            today_col_idx = 4 + (today - min_date).days
            today_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            for row in range(1, row_idx):
                cell = ws.cell(row=row, column=today_col_idx)
                # ヘッダー行は赤色で強調
                if row == 1:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True, size=9)

        # 全セルに罫線
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )

        for row in ws.iter_rows(min_row=1, max_row=row_idx-1, min_col=1, max_col=len(date_list)+3):
            for cell in row:
                cell.border = thin_border
                if cell.row > 1 and cell.column <= 3:
                    cell.alignment = Alignment(vertical="center")

    def _flatten_tasks(self, tasks: List[Task]) -> List[Task]:
        """タスクをフラット化（階層順に）"""
        result = []

        def add_recursive(task):
            result.append(task)
            for child in task.children:
                add_recursive(child)

        # 階層構造を構築
        for task in tasks:
            task.children = []
        task_dict = {t.id: t for t in tasks}
        root_tasks = []
        for task in tasks:
            if task.parent_id and task.parent_id in task_dict:
                task_dict[task.parent_id].add_child(task)
            elif task.parent_id is None:
                root_tasks.append(task)

        for task in root_tasks:
            add_recursive(task)

        return result

    def _get_task_level(self, task: Task, all_tasks: List[Task]) -> int:
        """タスクの階層レベルを取得"""
        level = 0
        task_dict = {t.id: t for t in all_tasks}
        current = task
        while current.parent_id and current.parent_id in task_dict:
            level += 1
            current = task_dict[current.parent_id]
        return level

    def _get_task_color_fill(self, task: Task) -> PatternFill:
        """タスクの進捗に応じた色を取得"""
        if task.is_milestone:
            return PatternFill(start_color="F44336", end_color="F44336", fill_type="solid")
        elif task.progress == 100:
            return PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        elif task.progress > 0:
            return PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
        else:
            return PatternFill(start_color="9E9E9E", end_color="9E9E9E", fill_type="solid")
